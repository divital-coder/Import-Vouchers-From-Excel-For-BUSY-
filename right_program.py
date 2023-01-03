import os
import openpyxl as excel
import pandas as pd

# mail modules


PATH = ""
FILENAME = ""
COLUMNS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
           "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
WARNING = """
\n\n
This program will only work if you put the target excel file in the hard coded directory inside the variable PATH,
not complying with it will result in errors and be carful while naming the target file after being asked by the prompt:)
\n\n
"""
print(WARNING)


def convert_csv_to_excel():
    global FILENAME
    if FILENAME.split(".")[1] == "csv":
        pandas_new = pd.read_csv(FILENAME)
        pandas_file = pd.ExcelWriter(FILENAME.split(".")[0] + ".xlsx")
        pandas_new.to_excel(pandas_file, index=False)

        pandas_file.save()

        FILENAME = FILENAME[0:-4] + ".xlsx"


def prepare_sales_sheet(wb):
    sales_sheet_creation = wb.create_sheet("SALES SHEET", 1)
    sales_sheet_creation.title = "sales_coded"
    sales_sheet_creation.sheet_properties.tabColor = "6DFC04"

    sales_sheet = wb[wb.sheetnames[1]]
    data_headers_sales = ["series"	, "invoice date"	, "invoice number",	"sale type", "party",	"ship to state", "ship to state",
                          "main store", "item name", "quantity", "unit", "invoice amount", "invoice amount", "hsn code", "transaction type"]
    for num in range(0, len(data_headers_sales)):
        sales_sheet[f"{COLUMNS[num]}1"].value = data_headers_sales[num]

    transaction_type_column = wb[wb.sheetnames[0]]["D"]
    transaction_type_list = []
    invoice_numbers_list = []
    invoice_date_list = []
    ship_to_state_list = []
    invoice_amount_list = []
    quantity_list = []
    for num in range(1, len(transaction_type_column)+1):
        if (wb[wb.sheetnames[0]][f"D{num}"].value == "Shipment" or wb[wb.sheetnames[0]][f"D{num}"].value == "Cancel"):
            transaction_type = wb[wb.sheetnames[0]][f"D{num}"].value
            transaction_type_list.append(transaction_type)

            invoice_number = wb[wb.sheetnames[0]][f"B{num}"].value
            invoice_numbers_list.append(invoice_number)

            invoice_date = wb[wb.sheetnames[0]][f"C{num}"].value
            invoice_date = str(invoice_date).split(" ")
            invoice_date_list.append(invoice_date[0])

            ship_to_state = wb[wb.sheetnames[0]][f"Y{num}"].value
            ship_to_state_list.append(ship_to_state)

            invoice_amount = wb[wb.sheetnames[0]][f"AB{num}"].value
            invoice_amount_list.append(invoice_amount)

            quantity = wb[wb.sheetnames[0]][f"J{num}"].value
            quantity_list.append(quantity)
    for num in range(2, len(transaction_type_list)+2):
        # invoice number colum
        sales_sheet[f"A{num}"].value = "MAIN"
        sales_sheet[f"B{num}"].value = invoice_date_list[num-2]
        sales_sheet[f"C{num}"].value = invoice_numbers_list[num-2]

        if(ship_to_state_list[num-2] == "DELHI"):
            sales_sheet[f"D{num}"].value = "Amazon Local"
        else:
            sales_sheet[f"D{num}"].value = "Amazon Igst"
        sales_sheet[f"E{num}"].value = "AMAZON SELLERS PVT. LTD."
        sales_sheet[f"F{num}"].value = ship_to_state_list[num-2]
        sales_sheet[f"G{num}"].value = ship_to_state_list[num-2]
        # transaction type column
        sales_sheet[f"H{num}"].value = "MAIN STORE"
        sales_sheet[f"I{num}"].value = "MULTIMEDIA SPEAKER SYSTEM"
        # sales_sheet[f"J{num}"].value=""
        sales_sheet[f"J{num}"].value = quantity_list[num-2]
        sales_sheet[f"K{num}"].value = "Pcs."
        sales_sheet[f"L{num}"].value = invoice_amount_list[num-2]
        sales_sheet[f"M{num}"].value = invoice_amount_list[num-2]
        sales_sheet[f"N{num}"].value = 8518
        sales_sheet[f"O{num}"].value = transaction_type_list[num-2]


def prepare_salesreturn_sheet(wb):
    salesreturn_sheet_creation = wb.create_sheet("SALES RETURN SHEET", 2)
    salesreturn_sheet_creation.title = "sales_return_coded"
    salesreturn_sheet_creation.sheet_properties.tabColor = "FC041B"

    sales_return_sheet = wb[wb.sheetnames[2]]
    data_headers_salesreturn = ["series"	, "credit note date", "credit note no", "sale type", "party",	"ship to state", "invoice number", "invoice date",
                                "invoice amount", "invoice amount", "ship to state", "main store", "item name", "quantity", "unit", "hsn code", "transaction type"]
    for num in range(0, len(data_headers_salesreturn)):
        sales_return_sheet[f"{COLUMNS[num]}1"].value = data_headers_salesreturn[num]

# ///////////////////////////////////////////////
    transaction_type_column = wb[wb.sheetnames[0]]["D"]
    transaction_type_list = []
    invoice_numbers_list = []
    invoice_date_list = []
    ship_to_state_list = []
    invoice_amount_list = []
    quantity_list = []
    credit_note_date_list = []
    credit_note_number_list = []
    for num in range(1, len(transaction_type_column)+1):
        if (wb[wb.sheetnames[0]][f"D{num}"].value == "Refund"):
            transaction_type = wb[wb.sheetnames[0]][f"D{num}"].value
            transaction_type_list.append(transaction_type)

            invoice_number = wb[wb.sheetnames[0]][f"B{num}"].value
            invoice_numbers_list.append(invoice_number)

            invoice_date = wb[wb.sheetnames[0]][f"C{num}"].value
            invoice_date = str(invoice_date).split(" ")
            invoice_date_list.append(invoice_date[0])

            ship_to_state = wb[wb.sheetnames[0]][f"Y{num}"].value
            ship_to_state_list.append(ship_to_state)

            invoice_amount = wb[wb.sheetnames[0]][f"AB{num}"].value
            invoice_amount_list.append(abs(invoice_amount))

            quantity = wb[wb.sheetnames[0]][f"J{num}"].value
            quantity_list.append(quantity)

            credit_note_date = wb[wb.sheetnames[0]][f"BZ{num}"].value
            credit_note_date = str(credit_note_date).split(" ")
            credit_note_date_list.append(credit_note_date[0])

            credit_note_number = wb[wb.sheetnames[0]][f"BY{num}"].value
            credit_note_number_list.append(credit_note_number)
    for num in range(2, len(transaction_type_list)+2):
        # invoice number colum
        sales_return_sheet[f"A{num}"].value = "MAIN"
        sales_return_sheet[f"B{num}"].value = credit_note_date_list[num-2]
        sales_return_sheet[f"C{num}"].value = credit_note_number_list[num-2]

        if(ship_to_state_list[num-2] == "DELHI"):
            sales_return_sheet[f"D{num}"].value = "Amazon Local"
        else:
            sales_return_sheet[f"D{num}"].value = "Amazon Igst"
        sales_return_sheet[f"E{num}"].value = "AMAZON SELLERS PVT. LTD."
        sales_return_sheet[f"F{num}"].value = ship_to_state_list[num-2]
        sales_return_sheet[f"G{num}"].value = invoice_numbers_list[num-2]
        sales_return_sheet[f"H{num}"].value = invoice_date_list[num-2]

        sales_return_sheet[f"I{num}"].value = invoice_amount_list[num-2]
        sales_return_sheet[f"J{num}"].value = invoice_amount_list[num-2]
        # transaction type column
        sales_return_sheet[f"K{num}"].value = ship_to_state_list[num-2]
        sales_return_sheet[f"L{num}"].value = "MAIN STORE"
        sales_return_sheet[f"M{num}"].value = "MULTIMEDIA SPEAKER SYSTEM"
        # sales_sheet[f"J{num}"].value=""
        sales_return_sheet[f"N{num}"].value = quantity_list[num-2]
        sales_return_sheet[f"O{num}"].value = "Pcs."

        sales_return_sheet[f"P{num}"].value = 8518
        sales_return_sheet[f"Q{num}"].value = transaction_type_list[num-2]


def access_excel_file():
    global FILENAME
    os_type = input(
        "Please enter the type of file system you are using : (linux/windows) : ").lower()
    if os_type == 'windows':
        PATH = "C:\\Users\\Pc\\OneDrive\\Desktop\\excel_sheets\\"
    elif os_type == "linux":
        PATH = r"/mnt/c/Users/Pc/onedrive/desktop/excel_sheets/"
    else:
        print("Run the program again! you made a typo above there!")
    os.chdir(PATH)
    FILENAME = input("Enter the name of the file : ")
    convert_csv_to_excel()
    workbook = excel.load_workbook(FILENAME)
    default_data_sheet = workbook[workbook.sheetnames[0]]
    prepare_sales_sheet(workbook)
    prepare_salesreturn_sheet(workbook)

    workbook.save(FILENAME)
    workbook.close()


access_excel_file()
